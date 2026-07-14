#!/usr/bin/env python3
"""
Migra el planning manual (xlsx con hojas "PANEL CONTROL MES") a Firestore-dev.

Reglas:
- Deduplica por ID: filas con el mismo ID en distintos meses son la misma
  pieza. `listoByMonth` se merge (TRUE en un mes si estaba TRUE en cualquier
  hoja de ese mes).
- Filas SIN ID se importan como piezas independientes (no arriesgamos).
- Filas con MISMO título pero DISTINTO ID son piezas distintas.
- El tipo se mapea a un blockId de la app teniendo en cuenta la variante
  (verde vs dorado) detectada por el color del cabecera de bloque en el xlsx.
- Tipos no reconocidos por la app (COLAS, DISTRIBUIDORES, PROMOS OTRAS…)
  se listan al final. NO se importan silenciosamente.

Uso:
    python3 scripts/import_manual_xlsx.py --dry-run
    python3 scripts/import_manual_xlsx.py --wipe --commit
"""

import argparse
import re
import sys
import time
import uuid
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials, firestore

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

XLSX_PATH = Path("/Volumes/MOVISTAR+/00 APPS/PANEL DE CONTROL/MATERIALES/PANEL DE CONTROL A REPLICAR EN JULIO.xlsx")
SERVICE_ACCOUNT_PATH = Path.home() / "Desktop" / "panel-control-movistar-dev-firebase-adminsdk-fbsvc-424a175db4.json"
PANEL_ID = "main"
DEFAULT_YEAR = 2026  # Ajustable si el xlsx representa otro año

MONTH_NUMBER = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# (TIPO_UPPER, variant) → blockId de la app
#   variant = "verde" (theme:9 en el xlsx) o "dorado" (theme:7).
#   Cuando la app solo tiene una variante para ese tipo, ambas caen en el mismo bloque.
TIPO_VARIANT_TO_BLOCK = {
    ("PROMO 20",         "verde"):  "block-1",
    ("PROMO 20",         "dorado"): "block-2",
    ("PROMO 40",         "verde"):  "block-3",
    ("PROMO 40",         "dorado"): "block-4",
    ("OTRAS DURACIONES", "verde"):  "block-5",
    ("OTRAS DURACIONES", "dorado"): "block-6",
    ("COMBO",            "verde"):  "block-7",
    ("COMBO",            "dorado"): "block-7",   # app solo tiene una variante
    ("BUMPER",           "verde"):  "block-8",
    ("BUMPER",           "dorado"): "block-9",
    ("ID",               "verde"):  "block-10",
    ("ID",               "dorado"): "block-10",
    ("PASOS A PUBLI",    "verde"):  "block-11",
    ("PASOS A PUBLI",    "dorado"): "block-12",
    ("PASO A PUBLI",     "verde"):  "block-11",
    ("PASO A PUBLI",     "dorado"): "block-12",
    ("INTRUSO",          "verde"):  "block-13",
    ("INTRUSO",          "dorado"): "block-13",  # app solo tiene una en la sección principal
    ("CANALES LALIGA",   "verde"):  "block-15",
    ("CANALES LALIGA",   "dorado"): "block-15",
    ("CANALES GOLF",     "verde"):  "block-16",
    ("CANALES GOLF",     "dorado"): "block-16",
    ("CAZA Y PESCA",     "verde"):  "block-17",
    ("CAZA Y PESCA",     "dorado"): "block-17",
    ("ARRANQUE",         "verde"):  "block-18",
    ("ARRANQUE",         "dorado"): "block-18",
    ("LOOP",             "verde"):  "block-14",  # Loop protección Pop-Ups (main)
    ("LOOP",             "dorado"): "block-19",  # Loop VOD
    ("PRE ROLL",         "verde"):  "block-20",
    ("PRE ROLL",         "dorado"): "block-21",
    # DISTRIBUIDORES es en realidad una LOOP (categoría a la que pertenecen).
    # El TIPO de la celda del xlsx debería decir "LOOP" pero por historia
    # dice "DISTRIBUIDORES". Lo mapeamos a Loop VOD (block-19) — encaja con
    # la posición de estos bloques en el xlsx original.
    ("DISTRIBUIDORES",   "verde"):  "block-19",
    ("DISTRIBUIDORES",   "dorado"): "block-19",
}

# Tipos del xlsx que NO tienen bloque equivalente en la app.
# Se listan al final del dry-run pero NO se importan a menos que aparezcan
# en TIPO_VARIANT_TO_BLOCK arriba. Estos tres son los que quedan fuera:
UNMAPPED_KNOWN_TIPOS = {"COLAS", "DISTRIBUIDORES", "PROMOS OTRAS"}


# ---------------------------------------------------------------------------
# PARSEO XLSX
# ---------------------------------------------------------------------------

def parse_month_from_sheet_name(name: str):
    m = re.match(r"^PANEL CONTROL (\w+)$", name.strip(), flags=re.IGNORECASE)
    if not m:
        return None
    return MONTH_NUMBER.get(m.group(1).upper())


def parse_date_string(txt: str, default_year: int):
    """
    Parsea una fecha de texto del xlsx aplicando default_year si el texto no
    incluye año. Devuelve un datetime o None.
    - '25/6'   → 25 junio default_year
    - '25/6/26' → 25 junio 2026 (yy en 2000-99 → 2000+yy)
    - '25/06/2026' → 25 junio 2026
    """
    if not txt:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d/%m", "%d-%m-%y", "%d-%m-%Y", "%d-%m"):
        try:
            d = datetime.strptime(txt, fmt)
            if fmt in ("%d/%m", "%d-%m"):
                d = d.replace(year=default_year)
            elif d.year < 100:
                d = d.replace(year=2000 + d.year)
            return d
        except ValueError:
            continue
    return None


def normalize_date_text(value, default_year: int):
    """Devuelve la fecha como 'DD/MM/YY' con año garantizado."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%y")
    txt = str(value).strip()
    if not txt:
        return ""
    d = parse_date_string(txt, default_year)
    if d:
        return d.strftime("%d/%m/%y")
    return txt


def date_text_to_iso(text: str, default_year: int):
    """Devuelve la fecha como 'YYYY-MM-DD' o cadena vacía."""
    if not text:
        return ""
    d = parse_date_string(text, default_year)
    if d:
        return d.strftime("%Y-%m-%d")
    return ""


def is_block_header_row(cell_titulo):
    """Cabecera de bloque: en col D pone "N SIMULTANEAS"."""
    if not cell_titulo:
        return False
    return "SIMULT" in str(cell_titulo).upper()


def block_header_variant(header_row_cell):
    """
    Devuelve 'verde', 'dorado' o None según el color de fondo del cabecera.
    theme:9 → verde ; theme:7 → dorado ; otro → None.
    """
    fill = header_row_cell.fill
    if not fill or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "theme":
        if fg.theme == 9:
            return "verde"
        if fg.theme == 7:
            return "dorado"
    if fg.type == "rgb" and fg.rgb:
        # Colores literales por si acaso
        if fg.rgb.upper().endswith("70AD47"):
            return "verde"
        if fg.rgb.upper().endswith("FFC000"):
            return "dorado"
    return None


def extract_rows(wb):
    """
    Recorre las hojas 'PANEL CONTROL *'. Para cada fila de datos, resuelve
    la variante (verde/dorado) mirando la última cabecera de bloque vista
    por ese TIPO en la misma hoja.

    Devuelve (rows, tipos_no_reconocidos, sheets_procesadas).
    """
    rows = []
    unknown_tipos = defaultdict(int)  # TIPO → cuántas filas
    sheets_procesadas = 0

    for sheet_name in wb.sheetnames:
        month_num = parse_month_from_sheet_name(sheet_name)
        if not month_num:
            continue
        sheets_procesadas += 1
        ws = wb[sheet_name]

        # variant_stack[TIPO] = variante del último bloque cabecera visto en ESTA hoja para ese TIPO
        current_variant_by_tipo = {}

        for r in range(3, ws.max_row + 1):
            listo_raw = ws.cell(row=r, column=1).value
            tipo_cell = ws.cell(row=r, column=3)
            titulo_cell = ws.cell(row=r, column=4)
            inicio = ws.cell(row=r, column=5).value
            fin = ws.cell(row=r, column=6).value
            rid = ws.cell(row=r, column=7).value

            tipo = str(tipo_cell.value or "").strip().upper()
            titulo = titulo_cell.value

            # Cabecera de bloque: actualizamos la variante actual para ese tipo
            if is_block_header_row(titulo):
                variant = block_header_variant(tipo_cell)
                if variant:
                    current_variant_by_tipo[tipo] = variant
                continue

            if titulo is None or str(titulo).strip() == "":
                continue

            # Fila de datos. Necesitamos (tipo, variante) para encontrar blockId.
            variant = current_variant_by_tipo.get(tipo, "verde")  # fallback: primera variante
            block_id = TIPO_VARIANT_TO_BLOCK.get((tipo, variant))
            if not block_id:
                unknown_tipos[tipo] += 1
                continue  # tipo no mapeado — se ignora

            rows.append({
                "sheet": sheet_name,
                "month": month_num,
                "year": DEFAULT_YEAR,
                "listo": bool(listo_raw) if isinstance(listo_raw, bool) else str(listo_raw or "").strip().lower() == "true",
                "tipo": tipo,
                "variant": variant,
                "block_id": block_id,
                "titulo": str(titulo).strip(),
                "inicio_text": normalize_date_text(inicio, DEFAULT_YEAR),
                "fin_text": normalize_date_text(fin, DEFAULT_YEAR),
                "id": str(rid or "").strip(),
            })
    return rows, unknown_tipos, sheets_procesadas


def dedupe(rows):
    """
    - Filas con ID: agrupar por ID. Primera aparición cronológica manda para
      metadatos. `listoByMonth` se agrega.
    - Filas sin ID: cada una es única.
    """
    grouped = defaultdict(list)
    unique_no_id = []
    for row in rows:
        if row["id"]:
            grouped[row["id"]].append(row)
        else:
            unique_no_id.append(row)

    pieces = []
    for rid, entries in grouped.items():
        entries.sort(key=lambda e: (e["year"], e["month"]))
        first = entries[0]
        listo_by_month = {}
        for e in entries:
            if e["listo"]:
                listo_by_month[f"{e['year']}-{str(e['month']).zfill(2)}"] = True
        pieces.append({
            "block_id": first["block_id"],
            "id": first["id"],
            "titulo": first["titulo"],
            "tipo": first["tipo"],
            "inicio_text": first["inicio_text"],
            "fin_text": first["fin_text"],
            "home_month": first["month"],
            "home_year": first["year"],
            "listo_by_month": listo_by_month,
        })

    for row in unique_no_id:
        listo_by_month = {}
        if row["listo"]:
            listo_by_month[f"{row['year']}-{str(row['month']).zfill(2)}"] = True
        pieces.append({
            "block_id": row["block_id"],
            "id": "",
            "titulo": row["titulo"],
            "tipo": row["tipo"],
            "inicio_text": row["inicio_text"],
            "fin_text": row["fin_text"],
            "home_month": row["month"],
            "home_year": row["year"],
            "listo_by_month": listo_by_month,
        })

    return pieces


# ---------------------------------------------------------------------------
# ESCRITURA A FIRESTORE
# ---------------------------------------------------------------------------

def build_row_key():
    return f"simp-{uuid.uuid4().hex[:12]}"


def wipe_existing(db):
    coll = db.collection("panels").document(PANEL_ID).collection("rows")
    docs = list(coll.stream())
    print(f"[wipe] borrando {len(docs)} docs existentes en panels/{PANEL_ID}/rows…")
    BATCH = 400
    for i in range(0, len(docs), BATCH):
        batch = db.batch()
        for d in docs[i:i + BATCH]:
            batch.delete(d.reference)
        batch.commit()
    print(f"[wipe] ✅ borrados {len(docs)}")


def write_pieces(db, pieces):
    coll = db.collection("panels").document(PANEL_ID).collection("rows")
    BATCH = 400
    written = 0
    now = firestore.SERVER_TIMESTAMP

    by_block = defaultdict(list)
    for p in pieces:
        by_block[p["block_id"]].append(p)

    for block_id, block_pieces in by_block.items():
        for i in range(0, len(block_pieces), BATCH):
            batch = db.batch()
            for order_index, p in enumerate(block_pieces[i:i + BATCH], start=i):
                row_key = build_row_key()
                inicio_iso = date_text_to_iso(p["inicio_text"], p["home_year"])
                fin_iso = date_text_to_iso(p["fin_text"], p["home_year"])
                listo_arr = list(p["listo_by_month"].keys())
                batch.set(coll.document(row_key), {
                    "blockId": block_id,
                    "title": p["titulo"],
                    "id": p["id"],
                    "genre": "",
                    "startDateText": p["inicio_text"],
                    "startDateISO": inicio_iso,
                    "endDateText": p["fin_text"],
                    "endDateISO": fin_iso,
                    "listoByMonth": listo_arr,
                    "actualizado": False,
                    "homeMonth": p["home_month"],
                    "homeYear": p["home_year"],
                    "orderIndex": order_index,
                    "createdAt": now,
                    "updatedAt": now,
                    "updatedBy": "manual-xlsx-import",
                    "deleted": False,
                })
                written += 1
            batch.commit()
        print(f"[write] block {block_id}: {len(block_pieces)} piezas")

    db.collection("panels").document(PANEL_ID).set({
        "createdAt": now, "schemaVersion": 1, "lastMigrationAt": now,
    }, merge=True)

    return written


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Solo cuenta, no toca Firestore")
    parser.add_argument("--wipe", action="store_true", help="Borra TODO en panels/main/rows antes de escribir")
    parser.add_argument("--commit", action="store_true", help="Confirma escritura real")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        print(f"❌ xlsx no encontrado: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"→ Leyendo {XLSX_PATH.name}…")
    wb = load_workbook(XLSX_PATH, data_only=False)
    rows, unknown_tipos, n_sheets = extract_rows(wb)
    print(f"  {len(rows)} filas mapeadas de {n_sheets} hojas de mes")

    if unknown_tipos:
        total_dropped = sum(unknown_tipos.values())
        print(f"\n⚠ Tipos NO mapeados en la app — {total_dropped} filas DESCARTADAS:")
        for tipo, n in sorted(unknown_tipos.items(), key=lambda kv: -kv[1]):
            print(f"    {tipo!r:25s} × {n}")
        print(f"  (para importarlas hay que decidir a qué bloque de la app van)")

    pieces = dedupe(rows)
    print(f"\n  {len(pieces)} piezas únicas tras deduplicar por ID")
    with_id = sum(1 for p in pieces if p["id"])
    print(f"    con ID (dedup real): {with_id}")
    print(f"    sin ID (importadas tal cual): {len(pieces) - with_id}")

    # Distribución por bloque
    print("\n  distribución por bloque:")
    per_block = defaultdict(int)
    for p in pieces:
        per_block[p["block_id"]] += 1
    for bid, n in sorted(per_block.items()):
        print(f"    {bid}: {n} piezas")

    if args.dry_run:
        print("\n[DRY-RUN] no escribo nada. Volver a ejecutar con --wipe --commit para escribir de verdad.")
        return

    if not args.commit:
        print("\nAborto: para escribir de verdad hace falta --commit (y opcionalmente --wipe).")
        return

    # -- REAL WRITE --
    if not SERVICE_ACCOUNT_PATH.exists():
        print(f"❌ service account no encontrado: {SERVICE_ACCOUNT_PATH}", file=sys.stderr)
        sys.exit(1)
    cred = credentials.Certificate(str(SERVICE_ACCOUNT_PATH))
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    if args.wipe:
        wipe_existing(db)

    print(f"\n→ Escribiendo {len(pieces)} piezas en Firestore-dev…")
    t0 = time.time()
    written = write_pieces(db, pieces)
    dt = time.time() - t0
    print(f"[commit] ✅ {written} filas escritas en {dt:.1f}s")


if __name__ == "__main__":
    main()
