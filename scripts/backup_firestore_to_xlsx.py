#!/usr/bin/env python3
"""
Backup automático: lee Firestore y construye un xlsx con el mismo formato
visual que el "DESCARGAR COPIA" de la app. Lo escribe en la carpeta que se
indique (por defecto ./backups-out/) para que un GitHub Action lo recoja
como artifact.

Diseñado para correr como GitHub Action cada 6h.

Variables de entorno esperadas:
    FIREBASE_SERVICE_ACCOUNT_JSON   contenido del JSON de service account
    PANEL_ID                        (opcional) por defecto "main"
    BACKUP_OUT_DIR                  (opcional) carpeta de salida; por defecto ./backups-out

Uso local (desarrollo):
    python3 scripts/backup_firestore_to_drive.py
"""

import calendar
import json
import os
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

PANEL_ID = os.environ.get("PANEL_ID", "main")

MONTH_NAMES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

# --- Definición de bloques del panel (calcada de createDefaultBlocks en app.js) --
# variant:  "verde" | "dorado" | "morado_oscuro" | "morado_claro"
BLOCKS = [
    {"id": "block-1",  "type": "PROMO 20",           "variant": "verde",  "max": 5},
    {"id": "block-2",  "type": "PROMO 20",           "variant": "dorado", "max": 5},
    {"id": "block-3",  "type": "PROMO 40",           "variant": "verde",  "max": 5},
    {"id": "block-4",  "type": "PROMO 40",           "variant": "dorado", "max": 5},
    {"id": "block-5",  "type": "OTRAS DURACIONES",   "variant": "verde",  "max": 5},
    {"id": "block-6",  "type": "OTRAS DURACIONES",   "variant": "dorado", "max": 5},
    {"id": "block-7",  "type": "COMBO",              "variant": "verde",  "max": 1},
    {"id": "block-8",  "type": "BUMPER",             "variant": "verde",  "max": 8},
    {"id": "block-9",  "type": "BUMPER",             "variant": "dorado", "max": 8},
    {"id": "block-10", "type": "ID",                 "variant": "verde",  "max": 1},
    {"id": "block-11", "type": "PASOS A PUBLI",      "variant": "verde",  "max": 5},
    {"id": "block-12", "type": "PASOS A PUBLI",      "variant": "dorado", "max": 5},
    {"id": "block-13", "type": "INTRUSO",            "variant": "verde",  "max": 10},
    {"id": "block-14", "type": "LOOP PROTECCION POP-UPS", "variant": "verde", "max": None},
    {"id": "block-15", "type": "CANALES LALIGA",     "variant": "dorado", "max": None},
    {"id": "block-16", "type": "CANALES GOLF",       "variant": "dorado", "max": None},
    {"id": "block-17", "type": "CANALES CAZA Y PESCA", "variant": "dorado", "max": None},
    {"id": "block-18", "type": "ARRANQUE",           "variant": "verde",  "max": 1},
    {"id": "block-19", "type": "LOOP",               "variant": "dorado", "max": 1},
    {"id": "block-20", "type": "PRE ROLL",           "variant": "verde",  "max": 5},
    {"id": "block-21", "type": "PRE ROLL",           "variant": "dorado", "max": 5},
    {"id": "block-22", "type": "OTRAS DURACIONES",   "variant": "morado_oscuro", "max": None},
    {"id": "block-23", "type": "BUMPER",             "variant": "morado_claro",  "max": None},
    {"id": "block-24", "type": "PASOS A PUBLI",      "variant": "morado_oscuro", "max": None},
    {"id": "block-25", "type": "INTRUSO",            "variant": "morado_claro",  "max": None},
    {"id": "block-26", "type": "PRE ROLL",           "variant": "morado_oscuro", "max": None},
    {"id": "block-27", "type": "PROMO 20",           "variant": "morado_oscuro", "max": None},
    {"id": "block-32", "type": "PROMO 40",           "variant": "morado_claro",  "max": None},
    {"id": "block-33", "type": "OTRAS DURACIONES",   "variant": "morado_oscuro", "max": None},
    {"id": "block-28", "type": "BUMPER",             "variant": "morado_claro",  "max": None},
    {"id": "block-29", "type": "PASOS A PUBLI",      "variant": "morado_oscuro", "max": None},
    {"id": "block-30", "type": "INTRUSO",            "variant": "morado_claro",  "max": None},
    {"id": "block-34", "type": "PRE ROLL",           "variant": "morado_oscuro", "max": None},
    {"id": "block-31", "type": "LOOP",               "variant": "morado_claro",  "max": None},
]

# Paleta (misma que el app.js buildExcelEdicionBuffer)
COLOR_MONTH_BAND = "BF8F00"
COLOR_BLUE_HDR   = "4472C4"
COLOR_GREEN      = "70AD47"
COLOR_GOLD       = "FFC000"
COLOR_PURPLE_D   = "AA87C6"
COLOR_PURPLE_L   = "C7A8E5"
COLOR_ORANGE_TXT = "FFC000"
COLOR_WEEKEND    = "ADACAC"
COLOR_DATA_TEXT  = "2E75B6"
COLOR_UPDATED    = "FF0000"
COLOR_DATE_ERROR = "FFC7CE"
COLOR_WHITE      = "FFFFFFFF"

VARIANT_TO_FILL = {
    "verde":         COLOR_GREEN,
    "dorado":        COLOR_GOLD,
    "morado_oscuro": COLOR_PURPLE_D,
    "morado_claro":  COLOR_PURPLE_L,
}

FIXED_COL_WIDTHS = [8, 9, 23, 90, 10, 9, 12]  # A LISTO, B MES, C TIPO, D TITULO, E INICIO VIG, F FIN VIG, G ID
DAY_COL_START = 8

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# LECTURA DE FIRESTORE
# ---------------------------------------------------------------------------

def read_all_rows(db):
    coll = db.collection("panels").document(PANEL_ID).collection("rows")
    rows = []
    for doc in coll.stream():
        data = doc.to_dict() or {}
        if data.get("deleted"):
            continue
        data["_rowKey"] = doc.id
        rows.append(data)
    return rows


# ---------------------------------------------------------------------------
# CÁLCULO DE MESES VISIBLES / RANGOS
# ---------------------------------------------------------------------------

def parse_iso(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def get_row_range(row):
    """Igual que getRowRange en app.js: si end < start o falta end, None."""
    end_d = parse_iso(row.get("endDateISO"))
    if not end_d:
        return None
    start_d = parse_iso(row.get("startDateISO"))
    if not start_d:
        # Sin inicio explícito: día 1 del homeMonth.
        hm = row.get("homeMonth")
        hy = row.get("homeYear")
        if isinstance(hm, int) and isinstance(hy, int):
            start_d = date(hy, hm, 1)
        else:
            return None
    if end_d < start_d:
        return None
    return start_d, end_d


def has_date_error(row):
    """Réplica de la validación de la app: rango inválido."""
    end_d = parse_iso(row.get("endDateISO"))
    start_d = parse_iso(row.get("startDateISO"))
    if start_d and end_d and end_d < start_d:
        return True
    return False


def months_touched(row):
    """Devuelve el conjunto {(year, month), ...} de meses en los que la fila
    debería aparecer: siempre su homeMonth, y todos los meses tocados por su
    rango si existe."""
    months = set()
    hm = row.get("homeMonth")
    hy = row.get("homeYear")
    if isinstance(hm, int) and isinstance(hy, int):
        months.add((hy, hm))
    rng = get_row_range(row)
    if rng:
        start_d, end_d = rng
        cur = date(start_d.year, start_d.month, 1)
        stop = date(end_d.year, end_d.month, 1)
        while cur <= stop:
            months.add((cur.year, cur.month))
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)
    return months


def block_daily_counts(rows_in_block, year, month, days):
    """Concurrencia por día del bloque en un mes concreto."""
    counts = [0] * (days + 1)  # 1-indexed
    month_start = date(year, month, 1)
    month_end = date(year, month, days)
    for row in rows_in_block:
        rng = get_row_range(row)
        if not rng:
            continue
        s, e = rng
        vs = max(s, month_start)
        ve = min(e, month_end)
        if ve < vs:
            continue
        for d in range(vs.day, ve.day + 1):
            counts[d] += 1
    return counts


def active_days_for_row(row, year, month, days):
    """Días del mes en los que la fila se emite."""
    rng = get_row_range(row)
    if not rng:
        return set()
    s, e = rng
    month_start = date(year, month, 1)
    month_end = date(year, month, days)
    vs = max(s, month_start)
    ve = min(e, month_end)
    if ve < vs:
        return set()
    return set(range(vs.day, ve.day + 1))


def get_weekend_days(year, month, days):
    """Sábado y domingo del mes (números de día)."""
    result = set()
    for d in range(1, days + 1):
        # weekday(): Monday=0, Sunday=6
        wd = date(year, month, d).weekday()
        if wd == 5 or wd == 6:
            result.add(d)
    return result


def normalize_listo_by_month(raw):
    """Firestore lo guarda como array de month keys. Devolvemos set."""
    if isinstance(raw, list):
        return set(raw)
    if isinstance(raw, dict):
        return set(k for k, v in raw.items() if v)
    return set()


# ---------------------------------------------------------------------------
# GENERACIÓN DEL XLSX
# ---------------------------------------------------------------------------

def build_xlsx(rows, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Agrupar filas por mes visible
    by_month = defaultdict(list)
    for row in rows:
        for (y, m) in months_touched(row):
            by_month[(y, m)].append(row)

    if not by_month:
        # Fichero vacío: al menos una hoja para que Excel no se queje
        ws = wb.create_sheet("VACIO")
        ws["A1"] = "No hay datos en Firestore."
        wb.save(output_path)
        return

    sorted_months = sorted(by_month.keys())
    now = datetime.now()
    now_key = (now.year, now.month)
    active_idx = sorted_months.index(now_key) if now_key in sorted_months else 0

    for year, month in sorted_months:
        _build_month_sheet(wb, year, month, by_month[(year, month)])

    # Activar la pestaña del mes actual al abrir
    wb.active = active_idx
    wb.save(output_path)


def _build_month_sheet(wb, year, month, month_rows):
    month_name = MONTH_NAMES_ES[month - 1].upper()
    days = calendar.monthrange(year, month)[1]
    weekends = get_weekend_days(year, month, days)

    ws = wb.create_sheet(f"PANEL CONTROL {month_name} {year}")

    # Anchos de columna
    for i, w in enumerate(FIXED_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for d in range(days):
        ws.column_dimensions[get_column_letter(DAY_COL_START + d)].width = 5

    # Fila 1: banda ocre con nombre del mes
    ws.merge_cells(start_row=1, start_column=DAY_COL_START,
                   end_row=1, end_column=DAY_COL_START + days - 1)
    title = ws.cell(row=1, column=DAY_COL_START)
    title.value = month_name
    title.font = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
    title.fill = PatternFill("solid", fgColor=COLOR_MONTH_BAND)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Fila 2: cabeceras
    fixed_headers = ["LISTO", "MES", "TIPO", "TITULO", "INICIO VIG", "FIN VIG", "ID"]
    for i, label in enumerate(fixed_headers, start=1):
        c = ws.cell(row=2, column=i, value=label)
        c.font = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", fgColor=COLOR_BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    for d in range(1, days + 1):
        c = ws.cell(row=2, column=DAY_COL_START + d - 1, value=d)
        c.font = Font(name="Calibri", size=13, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[2].height = 20

    # Bloques
    rows_by_block = defaultdict(list)
    for r in month_rows:
        rows_by_block[r.get("blockId")].append(r)

    current_row = 3
    for block in BLOCKS:
        block_rows = sorted(rows_by_block.get(block["id"], []),
                            key=lambda r: r.get("orderIndex", 0))
        block_fill = VARIANT_TO_FILL.get(block["variant"], COLOR_GREEN)

        data_count = max(1, len(block_rows))
        first_data_row = current_row + 1
        last_data_row = first_data_row + data_count - 1

        # Cabecera del bloque
        simul = f"{block['max']} SIMULTANEAS" if block["max"] else ""
        header_vals = ["", month_name, block["type"], simul, "", "", ""]
        for i, v in enumerate(header_vals, start=1):
            c = ws.cell(row=current_row, column=i, value=(v or None))
            is_title = (i == 4)
            c.font = Font(name="Calibri", size=11, bold=True,
                          color=(COLOR_ORANGE_TXT if is_title else COLOR_WHITE))
            c.fill = PatternFill("solid", fgColor=block_fill)
            c.alignment = Alignment(
                horizontal=("left" if i == 2 else "center"),
                vertical="center",
            )
            c.border = THIN_BORDER

        # Celdas de día del header: fórmulas SUM
        for d in range(1, days + 1):
            col_num = DAY_COL_START + d - 1
            col_letter = get_column_letter(col_num)
            c = ws.cell(row=current_row, column=col_num)
            c.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            c.font = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
            c.fill = PatternFill("solid", fgColor=block_fill)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER
        current_row += 1

        # Bloque sin piezas: fila vacía única
        if not block_rows:
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = THIN_BORDER
            for d in range(1, days + 1):
                c = ws.cell(row=current_row, column=DAY_COL_START + d - 1)
                c.border = THIN_BORDER
                if d in weekends:
                    c.fill = PatternFill("solid", fgColor=COLOR_WEEKEND)
            current_row += 1
            continue

        # Filas de datos
        for row in block_rows:
            month_key = f"{year}-{str(month).zfill(2)}"
            listo_set = normalize_listo_by_month(row.get("listoByMonth"))
            is_listo = month_key in listo_set

            row_vals = [
                "✓" if is_listo else "",
                month_name,
                block["type"],
                row.get("title") or "",
                row.get("startDateText") or "",
                row.get("endDateText") or "",
                row.get("id") or "",
            ]
            content_text_color = COLOR_UPDATED if row.get("actualizado") else COLOR_DATA_TEXT

            for i, v in enumerate(row_vals, start=1):
                c = ws.cell(row=current_row, column=i, value=v)
                if i == 1:
                    c.font = Font(name="Calibri", size=14, bold=True, color=COLOR_GREEN)
                else:
                    c.font = Font(name="Calibri", size=11, bold=True, color=content_text_color)
                c.alignment = Alignment(
                    horizontal=("left" if i == 4 else "center"),
                    vertical="center",
                )
                c.border = THIN_BORDER

            # Fechas como texto crudo
            ws.cell(row=current_row, column=5).number_format = "@"
            ws.cell(row=current_row, column=6).number_format = "@"

            # Resalte rojo si el rango es inválido
            if has_date_error(row):
                for col in (5, 6):
                    ws.cell(row=current_row, column=col).fill = PatternFill(
                        "solid", fgColor=COLOR_DATE_ERROR)

            # Celdas de día
            active = active_days_for_row(row, year, month, days)
            for d in range(1, days + 1):
                c = ws.cell(row=current_row, column=DAY_COL_START + d - 1)
                if d in active:
                    c.value = 1
                c.font = Font(name="Calibri", size=11)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = THIN_BORDER
                if d in weekends:
                    c.fill = PatternFill("solid", fgColor=COLOR_WEEKEND)

            current_row += 1


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def load_service_account():
    """De variable de entorno o de fichero local para desarrollo."""
    raw = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON")
    if raw:
        return json.loads(raw)
    fallback = Path.home() / "Desktop" / "panel-control-movistar-dev-firebase-adminsdk-fbsvc-424a175db4.json"
    if fallback.exists():
        return json.loads(fallback.read_text())
    raise SystemExit("❌ Falta service account (FIREBASE_SERVICE_ACCOUNT_JSON o fichero local)")


def main():
    sa_info = load_service_account()
    out_dir = Path(os.environ.get("BACKUP_OUT_DIR", "backups-out"))
    out_dir.mkdir(parents=True, exist_ok=True)

    # Init Firebase Admin
    firebase_admin.initialize_app(credentials.Certificate(sa_info))
    db = firestore.client()

    print("→ Leyendo filas de Firestore…")
    rows = read_all_rows(db)
    print(f"  {len(rows)} filas activas")

    now = datetime.now()
    filename = f"panel_backup_{now.strftime('%Y-%m-%d_%H%M')}.xlsx"
    out_path = out_dir / filename
    print(f"→ Generando {filename}…")
    build_xlsx(rows, str(out_path))
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✅ {out_path} — {size_kb:.1f} KB")
    print("[done]")


if __name__ == "__main__":
    main()
