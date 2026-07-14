#!/usr/bin/env python3
"""
Genera un xlsx con las filas del planning manual que la migración a la app
DESCARTARÍA porque su TIPO no tiene equivalente en los bloques de la app.

Tipos afectados: COLAS, DISTRIBUIDORES, o filas con TIPO vacío.

Cada fila del reporte incluye la hoja de origen y el motivo del descarte para
que los editores puedan corregir en el xlsx original o decidir a qué bloque
recategorizarlas.

Uso:
    python3 scripts/report_discarded_rows.py
"""

from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Reutilizamos las mismas rutas y reglas de la migración
import sys
sys.path.insert(0, str(Path(__file__).parent))
from import_manual_xlsx import (
    XLSX_PATH,
    DEFAULT_YEAR,
    TIPO_VARIANT_TO_BLOCK,
    parse_month_from_sheet_name,
    is_block_header_row,
    block_header_variant,
    normalize_date_text,
)

REPORT_PATH = XLSX_PATH.parent / "FILAS_DESCARTADAS_MIGRACION.xlsx"


def collect_discarded(wb):
    discarded = []
    for sheet_name in wb.sheetnames:
        month_num = parse_month_from_sheet_name(sheet_name)
        if not month_num:
            continue
        ws = wb[sheet_name]
        current_variant_by_tipo = {}
        for r in range(3, ws.max_row + 1):
            tipo_cell = ws.cell(row=r, column=3)
            titulo_cell = ws.cell(row=r, column=4)
            tipo = str(tipo_cell.value or "").strip().upper()
            titulo = titulo_cell.value

            if is_block_header_row(titulo):
                variant = block_header_variant(tipo_cell)
                if variant:
                    current_variant_by_tipo[tipo] = variant
                continue
            if titulo is None or str(titulo).strip() == "":
                continue

            variant = current_variant_by_tipo.get(tipo, "verde")
            block_id = TIPO_VARIANT_TO_BLOCK.get((tipo, variant))
            if block_id:
                continue  # esta fila SÍ se puede importar

            # COLAS es un formato descontinuado (confirmado por el equipo) —
            # se descarta silenciosamente, no lo incluimos en el reporte.
            if tipo == "COLAS":
                continue

            # Motivo del descarte
            if not tipo:
                motivo = "TIPO vacío en el xlsx"
            elif tipo in {"DISTRIBUIDORES", "PROMOS OTRAS"}:
                motivo = f"TIPO {tipo!r} no existe como bloque en la app"
            else:
                motivo = f"TIPO {tipo!r} sin mapeo en la app"

            listo = ws.cell(row=r, column=1).value
            inicio = ws.cell(row=r, column=5).value
            fin = ws.cell(row=r, column=6).value
            rid = ws.cell(row=r, column=7).value

            discarded.append({
                "hoja": sheet_name,
                "fila_xlsx": r,
                "listo": bool(listo) if isinstance(listo, bool) else str(listo or ""),
                "tipo": tipo,
                "titulo": str(titulo).strip(),
                "inicio": normalize_date_text(inicio, DEFAULT_YEAR),
                "fin": normalize_date_text(fin, DEFAULT_YEAR),
                "id": str(rid or "").strip(),
                "motivo": motivo,
            })
    return discarded


def write_report(discarded):
    wb = Workbook()
    ws = wb.active
    ws.title = "Filas descartadas"

    headers = ["HOJA", "FILA", "LISTO", "TIPO", "TITULO", "INICIO VIG", "FIN VIG", "ID", "MOTIVO DE DESCARTE"]
    widths  = [22,       7,      8,       18,     55,        12,           12,       14,     45]

    header_fill = PatternFill("solid", fgColor="FF4472C4")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (label, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    # Ordenar por hoja luego motivo luego título
    discarded.sort(key=lambda d: (d["hoja"], d["motivo"], d["titulo"]))

    for row_i, d in enumerate(discarded, start=2):
        vals = [d["hoja"], d["fila_xlsx"], d["listo"], d["tipo"], d["titulo"],
                d["inicio"], d["fin"], d["id"], d["motivo"]]
        for col_i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal=("left" if col_i in (5, 9) else "center"), vertical="center")

    ws.freeze_panes = "A2"
    wb.save(REPORT_PATH)


def main():
    if not XLSX_PATH.exists():
        print(f"❌ xlsx no encontrado: {XLSX_PATH}")
        return
    print(f"→ Leyendo {XLSX_PATH.name}…")
    wb = load_workbook(XLSX_PATH, data_only=False)
    discarded = collect_discarded(wb)
    print(f"  {len(discarded)} filas serían descartadas")

    from collections import Counter
    motivos = Counter(d["motivo"] for d in discarded)
    for m, n in motivos.most_common():
        print(f"    {n} × {m}")

    write_report(discarded)
    print(f"\n✅ Reporte guardado: {REPORT_PATH}")


if __name__ == "__main__":
    main()
